import openpyxl
from bs4 import BeautifulSoup

# Função para limpar HTML para texto mantendo a formatação
def limpar_html_para_texto_com_formatacao(html):
    # Analisa o HTML
    soup = BeautifulSoup(html, 'html.parser')

    # Obtém o texto mantendo a formatação
    texto_limp = ''
    for elem in soup.descendants:
        if elem.name == 'br':
            texto_limp += '\n'
        elif elem.name == 'p':
            texto_limp += '\n\n'
        elif elem.name is None and elem.strip():
            texto_limp += elem.strip() + ' '

    return texto_limp.strip()

# Caminho para a planilha original
caminho_planilha_original = './bling-pyton.xlsm'

# Carrega a planilha original
wb_original = openpyxl.load_workbook(caminho_planilha_original)
planilha_original = wb_original.active  # Você pode ajustar isso se a planilha não estiver na aba ativa

# Cria uma nova planilha
wb_nova = openpyxl.Workbook()
planilha_nova = wb_nova.active

# Copia o cabeçalho da planilha original para a nova planilha
for coluna in range(1, planilha_original.max_column + 1):
    planilha_nova.cell(row=1, column=coluna, value=planilha_original.cell(row=1, column=coluna).value)

# Itera sobre as linhas da planilha original e limpa o HTML na coluna B mantendo a formatação
for linha in range(2, planilha_original.max_row + 1):
    # Copia as células da coluna A até a coluna antes de B
    for coluna in range(1, 2):
        planilha_nova.cell(row=linha, column=coluna, value=planilha_original.cell(row=linha, column=coluna).value)
    
    celula_html = planilha_original.cell(row=linha, column=2)  # Coluna B
    texto_limpo = limpar_html_para_texto_com_formatacao(celula_html.value)
    planilha_nova.cell(row=linha, column=2, value=texto_limpo)  # Coloca o texto limpo na nova planilha

# Salva a nova planilha
caminho_planilha_nova = './planilha_nova.xlsx'
wb_nova.save(caminho_planilha_nova)